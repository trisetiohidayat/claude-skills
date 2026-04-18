#!/usr/bin/env python3
"""
Odoo Data Export to Excel
Export data from Odoo PostgreSQL database to Excel file.

Usage:
    python export_data.py --database <db_name> --model <model_name> --output <file.xlsx> [options]

Examples:
    # Export basic fields
    python export_data.py --database roedl_15_20260331 --model res_partner_bank \\
        --output ~/Downloads/export.xlsx --fields id,acc_number,partner_id,company_id

    # Export with external IDs (recommended for cross-database imports)
    python export_data.py --database roedl_15_20260331 --model res_partner_bank \\
        --output ~/Downloads/export.xlsx --fields acc_number,partner_id,company_id \\
        --include-external-id --related-as-external-id partner_id,company_id,bank_id

    # Export for Odoo-compatible import format (with id column as external ID)
    python export_data.py --database roedl_15_20260331 --model res.partner.bank \\
        --output ~/Downloads/export.xlsx --odoo-format
"""

import argparse
import csv
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser(description='Export Odoo data to Excel')
    parser.add_argument('--db-host', default='localhost', help='PostgreSQL host')
    parser.add_argument('--db-port', default='5432', help='PostgreSQL port')
    parser.add_argument('--db-user', default='odoo', help='PostgreSQL user')
    parser.add_argument('--db-password', default='odoo', help='PostgreSQL password')
    parser.add_argument('--database', required=True, help='Database name')
    parser.add_argument('--model', required=True, help='Model/Table name (e.g., res.partner.bank or res_partner_bank)')
    parser.add_argument('--output', required=True, help='Output file path (.xlsx or .csv)')
    parser.add_argument('--fields', default='*', help='Comma-separated fields to export (default: *)')
    parser.add_argument('--where', default='1=1', help='WHERE clause filter')
    parser.add_argument('--limit', type=int, default=0, help='Limit results (0 = no limit)')
    parser.add_argument('--include-external-id', action='store_true',
                        help='Add "id" column with external ID (module.name) for each record')
    parser.add_argument('--related-as-external-id', default='',
                        help='Comma-separated many2one fields to export as external IDs (e.g., partner_id,company_id)')
    parser.add_argument('--odoo-format', action='store_true',
                        help='Export in Odoo-compatible import format: id=external_id, many2one as /id columns')
    return parser.parse_args()


def get_table_name(model):
    """Convert Odoo model name to PostgreSQL table name."""
    return model.replace('.', '_')


def get_model_name(table_or_model):
    """Normalize to Odoo model name (with dots)."""
    return table_or_model.replace('_', '.') if '_' in table_or_model and '.' not in table_or_model else table_or_model


def get_connection(args):
    """Get PostgreSQL connection."""
    try:
        import psycopg2
    except ImportError:
        print("Installing psycopg2-binary...")
        os.system(f"{sys.executable} -m pip install psycopg2-binary")
        import psycopg2

    return psycopg2.connect(
        host=args.db_host,
        port=args.db_port,
        user=args.db_user,
        password=args.db_password,
        database=args.database
    )


def get_external_id(conn, model_name, record_id):
    """Get external ID (module.name) for a record from ir_model_data.

    Returns string like 'base.res_partner_1' or '__import__.res_partner_bank_22',
    or None if no external ID exists.
    """
    if record_id is None:
        return None
    cursor = conn.cursor()
    cursor.execute("""
        SELECT module, name
        FROM ir_model_data
        WHERE model = %s AND res_id = %s
        ORDER BY
            CASE WHEN module = 'base' THEN 0
                 WHEN module NOT LIKE '__%%' THEN 1
                 ELSE 2 END,
            id ASC
        LIMIT 1
    """, (model_name, record_id))
    row = cursor.fetchone()
    cursor.close()
    if row:
        return f"{row[0]}.{row[1]}"
    return None


def get_related_model(conn, table_name, field_name):
    """Try to determine the related model for a many2one field by looking at ir_model_fields."""
    cursor = conn.cursor()
    # Convert table name to Odoo model name
    model_name = table_name.replace('_', '.')
    cursor.execute("""
        SELECT relation
        FROM ir_model_fields
        WHERE model = %s AND name = %s AND ttype = 'many2one'
        LIMIT 1
    """, (model_name, field_name))
    row = cursor.fetchone()
    cursor.close()
    if row:
        return row[0]
    # Try with exact table name (some models don't follow convention)
    # Fallback: common patterns
    field_to_model = {
        'partner_id': 'res.partner',
        'company_id': 'res.company',
        'bank_id': 'res.bank',
        'currency_id': 'res.currency',
        'user_id': 'res.users',
        'product_id': 'product.product',
        'category_id': 'res.partner.category',
    }
    return field_to_model.get(field_name)


def fetch_data(conn, args):
    """Fetch data from PostgreSQL."""
    table_name = get_table_name(args.model)
    model_name = get_model_name(args.model)

    # Parse fields
    if args.fields == '*':
        fields_query = '*'
        fields_list = []
    else:
        fields_list = [f.strip() for f in args.fields.split(',')]
        # Ensure 'id' is included if we need external IDs
        if (args.include_external_id or args.odoo_format) and 'id' not in fields_list:
            fields_list = ['id'] + fields_list
        fields_query = ', '.join(fields_list)

    # Parse related fields for external ID export
    related_ext_id_fields = set()
    if args.related_as_external_id:
        related_ext_id_fields = {f.strip() for f in args.related_as_external_id.split(',')}
    if args.odoo_format:
        # In odoo-format mode, we'll convert many2one integer fields to external IDs
        # (all fields that end with _id and aren't 'id' itself)
        if fields_list:
            related_ext_id_fields |= {f for f in fields_list if f.endswith('_id') and f != 'id'}

    # Build query
    query = f'SELECT {fields_query} FROM {table_name} WHERE {args.where}'
    if args.limit > 0:
        query += f' LIMIT {args.limit}'

    cursor = conn.cursor()
    cursor.execute(query)

    if args.fields == '*':
        fields_list = [desc[0] for desc in cursor.description]
        if args.odoo_format:
            related_ext_id_fields = {f for f in fields_list if f.endswith('_id') and f != 'id'}
    else:
        fields_list = [f.strip() for f in args.fields.split(',')]
        if (args.include_external_id or args.odoo_format) and 'id' not in fields_list:
            fields_list = ['id'] + [f.strip() for f in args.fields.split(',')]

    rows = cursor.fetchall()
    cursor.close()

    # Build output columns
    # In odoo-format: id → external_id, field_id → field_id/id (external ID of related)
    if args.include_external_id or args.odoo_format or related_ext_id_fields:
        id_col_idx = fields_list.index('id') if 'id' in fields_list else None
        rows, fields_list = enrich_with_external_ids(
            conn, model_name, table_name, rows, fields_list,
            related_ext_id_fields,
            include_ext_id=(args.include_external_id or args.odoo_format),
            odoo_format=args.odoo_format
        )

    return fields_list, rows


def enrich_with_external_ids(conn, model_name, table_name, rows, fields_list,
                              related_ext_id_fields, include_ext_id=False, odoo_format=False):
    """Add external ID columns to the result set.

    - If include_ext_id: replace 'id' column value with external ID string
    - For each field in related_ext_id_fields: add a '<field>/id' column with the related record's external ID
    """
    id_col_idx = fields_list.index('id') if 'id' in fields_list else None

    # Pre-build related model map
    related_models = {}
    for field in related_ext_id_fields:
        related_models[field] = get_related_model(conn, table_name, field)

    # Build new field list
    new_fields = []
    for f in fields_list:
        if f == 'id' and odoo_format:
            new_fields.append('id')  # Will contain external ID
        elif f in related_ext_id_fields and odoo_format:
            # Replace field with field/id (external ID format)
            new_fields.append(f'{f}/id')
        else:
            new_fields.append(f)
            if f in related_ext_id_fields and not odoo_format:
                # In non-odoo-format: keep original column AND add /id column
                new_fields.append(f'{f}/id')

    # Process rows
    # Cache external ID lookups to avoid repeated queries
    ext_id_cache = {}  # (model, record_id) -> external_id

    def lookup_ext_id(model, record_id):
        if model is None or record_id is None:
            return None
        key = (model, record_id)
        if key not in ext_id_cache:
            ext_id_cache[key] = get_external_id(conn, model, record_id)
        return ext_id_cache[key]

    new_rows = []
    print(f"  Enriching {len(rows)} rows with external IDs...")
    for i, row in enumerate(rows):
        row = list(row)
        new_row = []
        for col_idx, (field, val) in enumerate(zip(fields_list, row)):
            if field == 'id' and (include_ext_id or odoo_format):
                # Replace DB id with external ID
                ext_id = lookup_ext_id(model_name, val)
                new_row.append(ext_id if ext_id else f'__import__.{table_name}_{val}')
            elif field in related_ext_id_fields:
                if odoo_format:
                    # Replace with external ID of related record
                    related_model = related_models.get(field)
                    ext_id = lookup_ext_id(related_model, val) if val else None
                    new_row.append(ext_id)
                else:
                    # Keep original value
                    new_row.append(val)
                    # Then add external ID column
                    related_model = related_models.get(field)
                    ext_id = lookup_ext_id(related_model, val) if val else None
                    new_row.append(ext_id)
            else:
                new_row.append(val)

        new_rows.append(new_row)
        if (i + 1) % 50 == 0:
            print(f"  Processed {i + 1}/{len(rows)} rows...")

    return new_rows, new_fields


def write_excel(fields, rows, output_path):
    """Write data to Excel file with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ext_id_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for ext ID cols
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Write header
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        # Highlight /id columns differently
        if field == 'id' or field.endswith('/id'):
            cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        else:
            cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            if value is None:
                display_value = ''
            else:
                display_value = value if not isinstance(value, (int, float)) else value

            cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            # Highlight external ID columns
            if fields[col_idx - 1] == 'id' or fields[col_idx - 1].endswith('/id'):
                cell.fill = ext_id_fill

    # Auto-adjust column widths
    for col_idx, field in enumerate(fields, 1):
        max_length = len(str(field))
        for row in rows:
            val = row[col_idx - 1]
            if val is not None:
                max_length = max(max_length, len(str(val)))
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(output_path)
    return len(rows)


def write_csv(fields, rows, output_path):
    """Write data to CSV file."""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for row in rows:
            row_data = ['' if val is None else str(val) for val in row]
            writer.writerow(row_data)
    return len(rows)


def main():
    args = parse_args()

    print(f"Connecting to: {args.database} ({args.db_host}:{args.db_port})")
    print(f"Model: {args.model}")
    print(f"Table: {get_table_name(args.model)}")
    print(f"Fields: {args.fields}")
    if args.include_external_id:
        print("Mode: Include External IDs (id column = module.name)")
    if args.odoo_format:
        print("Mode: Odoo-compatible format (id=external_id, many2one as /id)")
    if args.related_as_external_id:
        print(f"Related fields as external ID: {args.related_as_external_id}")
    print("-" * 50)

    try:
        conn = get_connection(args)
        fields, rows = fetch_data(conn, args)
        conn.close()

        print(f"Fetched {len(rows)} records")
        print(f"Fields: {', '.join(fields)}")

        output_path = os.path.expanduser(args.output)
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

        if output_path.endswith('.xlsx') or output_path.endswith('.xls'):
            count = write_excel(fields, rows, output_path)
            print(f"\nExcel file saved: {output_path}")
        else:
            count = write_csv(fields, rows, output_path)
            print(f"\nCSV file saved: {output_path}")

        print(f"Total records exported: {count}")

    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
